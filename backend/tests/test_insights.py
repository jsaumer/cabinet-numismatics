"""Phase 4: breakdowns, per-item gains, Excel export."""

import io

from openpyxl import load_workbook

from tests.conftest import COIN


def _create(client, payload):
    resp = client.post("/api/items", json=payload)
    assert resp.status_code == 201
    return resp.json()


def _grade_id(client, code):
    grades = client.get("/api/grades", params={"scale": "sheldon"}).json()
    return next(g["id"] for g in grades if g["code"] == code)


def test_breakdowns(client):
    gid = _grade_id(client, "MS-63")
    a = _create(
        client,
        {
            **COIN,
            "year": 1932,
            "acquisition_price": 100.0,
            "acquisition_date": "2024-05-01",
            "grade_id": gid,
            "tags": ["silver"],
        },
    )
    client.post(f"/api/items/{a['id']}/estimates", json={"estimated_value": 150.0})

    _create(
        client,
        {
            **COIN,
            "type": "note",
            "country": "Canada",
            "year": 1954,
            "acquisition_price": 20.0,
            "acquisition_date": "2024-08-01",
            "tags": ["silver", "notes"],
        },
    )

    # sold and wishlist items don't appear in breakdowns
    sold = _create(client, {**COIN, "acquisition_price": 10.0})
    client.patch(f"/api/items/{sold['id']}", json={"status": "sold", "sold_price": 15.0})
    _create(client, {**COIN, "status": "wishlist"})

    body = client.get("/api/stats/breakdowns").json()

    countries = {e["key"]: e for e in body["by_country"]}
    assert set(countries) == {"United States", "Canada"}
    assert countries["United States"]["count"] == 1
    assert countries["United States"]["estimated_value"] == 150.0
    assert countries["Canada"]["cost_basis"] == 20.0

    assert {e["key"] for e in body["by_type"]} == {"coin", "note"}
    assert [e["key"] for e in body["by_decade"]] == ["1930s", "1950s"]

    grades = {e["key"] for e in body["by_grade"]}
    assert grades == {"MS-63", "ungraded"}

    tags = {e["key"]: e["count"] for e in body["by_tag"]}
    assert tags == {"silver": 2, "notes": 1}  # multi-tag items count in each

    years = {e["key"]: e for e in body["acquisitions_by_year"]}
    assert years["2024"]["count"] == 2 and years["2024"]["cost_basis"] == 120.0


def test_breakdowns_currency_rule(client):
    _create(client, {**COIN, "acquisition_price": 50.0, "currency": "CAD"})
    body = client.get("/api/stats/breakdowns").json()
    us = next(e for e in body["by_country"] if e["key"] == "United States")
    assert us["count"] == 1  # counted
    assert us["cost_basis"] == 0.0  # but CAD money not summed into USD


def test_gains(client):
    winner = _create(client, {**COIN, "acquisition_price": 100.0})
    client.post(f"/api/items/{winner['id']}/estimates", json={"estimated_value": 180.0})
    loser = _create(client, {**COIN, "denomination": "5 cents", "acquisition_price": 60.0})
    client.post(f"/api/items/{loser['id']}/estimates", json={"estimated_value": 40.0})

    sold = _create(client, {**COIN, "acquisition_price": 80.0})
    client.patch(
        f"/api/items/{sold['id']}",
        json={"status": "sold", "sold_date": "2026-07-01", "sold_price": 130.0},
    )

    _create(client, COIN)  # no estimate: absent from gains

    body = client.get("/api/stats/gains").json()
    assert [e["gain"] for e in body["unrealized"]] == [80.0, -20.0]  # sorted desc
    assert body["unrealized"][0]["label"].startswith("United States 25 cents 1932")
    assert len(body["realized"]) == 1
    assert body["realized"][0]["gain"] == 50.0
    assert body["realized"][0]["value"] == 130.0


def test_xlsx_export(client, coin):
    client.post(f"/api/items/{coin['id']}/estimates", json={"estimated_value": 150.0})

    resp = client.get("/api/items/export.xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[0] == "id" and "latest_value" in headers
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = dict(zip(headers, rows[0], strict=True))
    assert row["country"] == "United States"
    assert float(row["latest_value"]) == 150.0
